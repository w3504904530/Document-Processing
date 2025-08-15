#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件处理工具程序
支持CSV、Excel、MySQL的读取和智能对比功能
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import module.files


# 以df1的A列为基准，将df2的A列与df1的A列进行匹配，如果匹配到，则将df2的B列、C列、D列的值赋值给df1的B列、C列、D列    
def merge_and_reorder(df1, df2, comparison_column, preserve_order_by=None, column_sort_strategy='alternating'):
	"""将两个 DataFrame 合并并按照列交替排列
	参数:
	  preserve_order_by: None | 'df1' | 'df2'  如果指定 'df2' 则结果会按 df2 的行顺序保留（优先保留 df2 的顺序）
	  column_sort_strategy: 'alternating' | 'grouped' | 'alphabetical'  列排序策略
		- 'alternating': 交替排列（默认）
		- 'grouped': 分组排列（先df1的所有列，再df2的所有列）
		- 'alphabetical': 按字母顺序排列
	"""
	# 保留原始列集合用于后续列名构建
	df1_cols = list(df1.columns)
	df2_cols = list(df2.columns)

	# 根据是否需要保留某一侧的顺序选择合并方式
	if preserve_order_by == 'df2':
		# 以 df2 的顺序为主，用 left merge 保留 df2 顺序；设置后缀使得原 df1 列标记为 _1
		merged_df = pd.merge(df2, df1, on=comparison_column, how='left', suffixes=('', '_1'))
	elif preserve_order_by == 'df1':
		merged_df = pd.merge(df1, df2, on=comparison_column, how='left', suffixes=('', '_2'))
	else:
		# 默认行为：outer 合并（原始实现）
		merged_df = pd.merge(df1, df2, on=comparison_column, how='outer', suffixes=('', '_2'))


	# 优化后的列名顺序重构
	# 找出公共列（不包含比较列）
	common_columns = [col for col in df1_cols if col in df2_cols and col != comparison_column]
	
	# 根据排序策略构建列顺序
	if column_sort_strategy == 'alternating':
		# 交替排列策略
		final_columns = [comparison_column]
		
		# 使用更清晰的逻辑构建交替列
		for col in common_columns:
			if preserve_order_by == 'df2':
				# df2优先时：df1列名加后缀，df2列名保持原样
				final_columns.extend([f"{col}_1", col])
			else:
				# df1优先或默认时：df1列名保持原样，df2列名加后缀
				final_columns.extend([col, f"{col}_2"])
		
		# 优化独有列的处理
		df1_unique_cols = [col for col in df1_cols if col not in df2_cols and col != comparison_column]
		df2_unique_cols = [col for col in df2_cols if col not in df1_cols and col != comparison_column]
		
		# 根据preserve_order_by决定独有列的顺序
		if preserve_order_by == 'df2':
			remaining_columns = df2_unique_cols + df1_unique_cols
		else:
			remaining_columns = df1_unique_cols + df2_unique_cols
			
		final_columns.extend(remaining_columns)
		
	elif column_sort_strategy == 'grouped':
		# 分组排列策略：先df1的所有列，再df2的所有列
		final_columns = [comparison_column]
		
		# 添加df1的所有列（包括公共列和独有列）
		for col in df1_cols:
			if col != comparison_column:
				if col in common_columns:
					final_columns.append(col)  # 公共列保持原列名
				else:
					final_columns.append(col)  # df1独有列
		
		# 添加df2的所有列（公共列加后缀，独有列保持原列名）
		for col in df2_cols:
			if col != comparison_column:
				if col in common_columns:
					final_columns.append(f"{col}_2")  # 公共列加后缀
				else:
					final_columns.append(col)  # df2独有列
					
	elif column_sort_strategy == 'alphabetical':
		# 按字母顺序排列
		final_columns = sorted(merged_df.columns)
	
	# 确保所有列都存在（处理合并后的列名变化）
	valid_columns = []
	for col in final_columns:
		if col in merged_df.columns:
			valid_columns.append(col)
	
	# 添加任何遗漏的列
	missing_columns = [col for col in merged_df.columns if col not in valid_columns]
	valid_columns.extend(missing_columns)
	
	final_columns = valid_columns
	
	# 按新顺序重新排列
	# 统一生成用于高亮的成对列名（自动匹配同名列）
	column_pairs = []
	suffix_other = '_1' if preserve_order_by == 'df2' else '_2'
	for base_col in common_columns:
		left_name = f"{base_col}" if suffix_other == '_2' else f"{base_col}{suffix_other}"
		right_name = f"{base_col}{suffix_other}" if suffix_other == '_2' else f"{base_col}"
		# 确保这两列存在于结果中
		if left_name in merged_df.columns and right_name in merged_df.columns:
			column_pairs.append((left_name, right_name))
	
	return merged_df[final_columns], column_pairs


def highlight_differences(output_path, column_pairs):
	"""在 Excel 中高亮显示不同的值"""
	try:
		# 打开 Excel 文件
		workbook = load_workbook(output_path)
		sheet = workbook.active

		# 设置高亮样式（黄色填充）
		fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

		# 优化：预先构建列名到列索引的映射
		header_row = sheet[1]
		col_mapping = {}
		for cell in header_row:
			if cell.value:
				col_mapping[cell.value] = cell.col_idx

		# 优化：只处理有效的列对（自动匹配同名列形成的对）
		valid_column_pairs = []
		for col1, col2 in column_pairs:
			if col1 in col_mapping and col2 in col_mapping:
				valid_column_pairs.append((col_mapping[col1], col_mapping[col2]))

		# 批量处理高亮
		max_row = sheet.max_row
		highlighted_cells = 0
		
		for col1_index, col2_index in valid_column_pairs:
			for row in range(2, max_row + 1):  # 跳过标题行
				value1 = sheet.cell(row=row, column=col1_index).value
				value2 = sheet.cell(row=row, column=col2_index).value

				# 如果值不同，高亮显示
				if value1 != value2:
					sheet.cell(row=row, column=col1_index).fill = fill
					sheet.cell(row=row, column=col2_index).fill = fill
					highlighted_cells += 2

		# 保存文件
		workbook.save(output_path)
		print(f"已完成高亮处理: {output_path}，共高亮了 {highlighted_cells} 个单元格")
	except Exception as e:
		print(f"高亮显示出错: {e}")


def save_to_excel(df, output_path):
	"""保存 DataFrame 到 Excel 文件"""
	try:
		df.to_excel(output_path, index=False)
		print(f"拼接结果已保存到: {output_path}")
	except Exception as e:
		print(f"保存 Excel 文件时出错: {e}")


def data_comparison(col, file1, file2, preserve_order_by, column_sort_strategy, output_path):
	"""对比两个文件并输出拼接结果到 Excel，并高亮显示不同
	参数:
	  col: 比较列名
	  file1, file2: 文件路径
	  preserve_order_by: None | 'df1' | 'df2'  行顺序保留策略
	  column_sort_strategy: 'alternating' | 'grouped' | 'alphabetical'  列排序策略
	  output_path: 输出Excel路径
	"""

	df1 = module.files.read_file(file1)
	df2 = module.files.read_file(file2)

	if df1 is None or df2 is None:
		print("文件读取失败，程序终止。")
		return

	print(f"df1 列名: {df1.columns.tolist()}")
	print(f"df2 列名: {df2.columns.tolist()}")
	print(f"使用列排序策略: {column_sort_strategy}")

	# 合并数据框并按列交替排列
	merged_df, column_pairs = merge_and_reorder(df1, df2, col, preserve_order_by, column_sort_strategy)

	# 保存到 Excel
	save_to_excel(merged_df, output_path)

	# 所有策略均支持自动匹配同名列并成对高亮
	if column_pairs:
		highlight_differences(output_path, column_pairs)
	else:
		print("未找到可高亮的成对列。")


if __name__ == "__main__":
	file1 = r'C:\Users\35049\OneDrive\PYthon\Document-Processing\test\1.xlsx'
	file2 = r'C:\Users\35049\OneDrive\PYthon\Document-Processing\test\2.csv'
	# 指定比较的列名
	comparison_column = 'A'
	
	# 演示不同的列排序策略（输出到不同文件，避免覆盖）
	print("=== 交替排列策略 ===")
	data_comparison(comparison_column, file1, file2, preserve_order_by='Any', column_sort_strategy='alternating', output_path='./data/out_alternating.xlsx')
	
	# print("\n=== 分组排列策略 ===")
	# data_comparison(comparison_column, file1, file2, preserve_order_by='Any', column_sort_strategy='grouped', output_path='./data/out_grouped.xlsx')
	
	# print("\n=== 字母顺序排列策略 ===")
	# data_comparison(comparison_column, file1, file2, preserve_order_by='Any', column_sort_strategy='alphabetical', output_path='./data/out_alphabetical.xlsx')

